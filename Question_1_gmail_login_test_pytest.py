
#**********************************************************************************************
#run this code with
#pytest -v -s Question_1_gmail_login_test_pytest.py --alluredir=allure_reports
#**********************************************************************************************

##Question 1:
#Write UI a test case for the below scenario
# 1. Create an csv/excel sheet with username, password , expected results columns with values
# 2. Write a scenario to load the gmail.com
# 3. Perform login operation with expected validations based on the excel sheet(Data drive it)
# 4. Generate html/json report based on the pytest/behave/robot framework configuration



import pytest 
import allure # for reports
from selenium import webdriver 
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager

import openpyxl
from time import sleep


@pytest.fixture(params=["chrome","firefox","edge"],scope="class")
def driver_setup(request):
    global driver
    if request.param == "chrome":
        driver = webdriver.Chrome(executable_path = ChromeDriverManager().install())
    if request.param == "firefox":
       driver = webdriver.Firefox(executable_path = GeckoDriverManager().install())

    if request.param == "edge":
       driver = webdriver.Edge(executable_path = EdgeChromiumDriverManager().install())


    driver.maximize_window()
    yield
    driver.close()
     
 

##reading user name and password from excel file
def read_excel_data():
    userlist = []
    path = r"D:\test_gmail_login.xlsx"

    workbook = openpyxl.load_workbook(path)

    sheet = workbook.get_sheet_by_name("Sheet1")

    rows = sheet.max_row + 1

    for r in range(2, rows):
        username = sheet.cell(r, 1).value
        password = sheet.cell(r, 2).value

        tuple = (username,password)

        userlist.append(tuple)
    #print(userlist)

    return userlist



@allure.description("Automatin gmail.com login with different username and passwords  ")
@allure.severity(severity_level = "CRITICAL")
@pytest.mark.parametrize("username , password", read_excel_data())
def test_login(driver_setup, username, password):
    try :
        driver.get("https://accounts.google.com/signin/v2/identifier?continue=https%3A%2F%2Fmail.google.com%2Fmail%2F&service=mail&sacu=1&rip=1&flowName=GlifWebSignIn&flowEntry=ServiceLogin")
        #driver.get("https://orangehrm-demo-6x.orangehrmlive.com")
        driver.implicitly_wait(15)
        loginBox = driver.find_element(By.XPATH, '//*[@id ="identifierId"]')
        loginBox.send_keys(username)
    
        nextButton = driver.find_element(By.XPATH, '//*[@id ="identifierNext"]')
        nextButton[0].click()
    
        passWordBox = driver.find_element(By.XPATH,'//*[@id ="password"]/div[1]/div / div[1]/input')
        passWordBox.send_keys(password)
    
        nextButton = driver.find_elements(By.XPATH, '//*[@id ="passwordNext"]')
        nextButton[0].click()
    
        print('Login Successful...!!')
        assert "You\'re signed in" in driver.find_element(By.className, '//*[@class="You\'re signed in"]')
    finally:
        if(AssertionError):
            allure.attach(driver.get_screenshot_as_png(), name = "Invalid Credentials", attachment_type = allure.attachment_type.PNG)
        print('Login Failed')
    sleep(5)


